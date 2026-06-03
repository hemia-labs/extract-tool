from pathlib import Path

from openpyxl import load_workbook

from app.extractors.base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    def extract(self, path: Path, ocr: str, language: str, output: str) -> ExtractionResult:
        workbook = load_workbook(path, read_only=True, data_only=True)
        pages = []

        for index, sheet in enumerate(workbook.worksheets, start=1):
            rows = [
                ["" if cell is None else str(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]
            text = self._to_markdown(sheet.title, rows) if output == "markdown" else self._to_text(sheet.title, rows)
            pages.append({"page": index, "text": text, "confidence": None})

        joined = "\n\n".join(page["text"] for page in pages)
        return {
            "text": joined,
            "pages": pages,
            "metadata": {"ocrUsed": False, "pages": len(pages), "characters": len(joined)},
        }

    @staticmethod
    def _to_text(title: str, rows: list[list[str]]) -> str:
        body = "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)
        return f"Sheet: {title}\n{body}".strip()

    @staticmethod
    def _to_markdown(title: str, rows: list[list[str]]) -> str:
        if not rows:
            return f"## Hoja: {title}"
        width = max(len(row) for row in rows)
        padded = [row + [""] * (width - len(row)) for row in rows]
        header = "| " + " | ".join(padded[0]) + " |"
        sep = "| " + " | ".join("---" for _ in range(width)) + " |"
        body = ["| " + " | ".join(row) + " |" for row in padded[1:]]
        return "\n".join([f"## Hoja: {title}", "", header, sep, *body])
